import json
import openpyxl
import logging
from pathlib import Path
from typing import List, Dict

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


class SectionValidator:
    """Class to validate TOC and Spec JSONL files and write XLSX validation report."""

    def __init__(self, toc_file: Path, spec_file: Path, output_file: Path):
        self.toc_file = toc_file
        self.spec_file = spec_file
        self.output_file = output_file
        self.toc_entries = []
        self.spec_entries = []

    def load_jsonl(self, path: Path) -> List[Dict]:
        try:
            with path.open(encoding="utf-8") as f:
                return [json.loads(line) for line in f if line.strip()]
        except Exception as e:
            logging.error(f"Error loading JSONL from {path}: {e}")
            return []

    def validate(self):
        """
        Load TOC and Spec JSONL files and verify that they contain data.
        Returns: bool - True if both TOC and Spec entries are loaded successfully and non-empty,
                        False otherwise.
        """
        self.toc_entries = self.load_jsonl(self.toc_file)
        self.spec_entries = self.load_jsonl(self.spec_file)

        if not self.toc_entries or not self.spec_entries:
            logging.error("Empty TOC or Spec input; cannot generate validation report.")
            return False

        return True

    def write_report(self):
        """
        Create and save an Excel workbook summarizing validation results:
        - Counts of TOC vs Spec sections and their overlap.
        - Missing sections in Spec compared to TOC.
        - Extra sections in Spec not found in TOC, with tags.
        - Tag frequency summary across Spec sections.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validation"

        toc_ids = [t["section_id"] for t in self.toc_entries]
        spec_ids = [s["section_id"] for s in self.spec_entries]
        spec_tags_map = {s["section_id"]: s.get("tags", []) for s in self.spec_entries}

        missing_in_spec = sorted(set(toc_ids) - set(spec_ids))
        extra_in_spec = sorted(set(spec_ids) - set(toc_ids))

        # Summary section
        ws.append(["Check", "TOC Count", "Parsed Count", "Missing Count", "Extra Count"])
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        ws.append([
            "Section Count",
            len(toc_ids),
            len(spec_ids),
            len(missing_in_spec),
            len(extra_in_spec)
        ])

        ws.append([])

        # Missing sections
        ws.append(["Missing Sections in Parsed"])
        ws.append(["Section ID", "Title"])
        for cell in ws[ws.max_row]:
            cell.font = openpyxl.styles.Font(bold=True)
        for sec_id in missing_in_spec:
            title = next((t["title"] for t in self.toc_entries if t["section_id"] == sec_id), "")
            ws.append([sec_id, title])
        if not missing_in_spec:
            ws.append(["None"])

        ws.append([])

        # Extra sections
        ws.append(["Extra Sections in Parsed (Not in TOC)"])
        ws.append(["Section ID", "Title", "Tags"])
        for cell in ws[ws.max_row]:
            cell.font = openpyxl.styles.Font(bold=True)
        for sec_id in extra_in_spec:
            title = next((s["title"] for s in self.spec_entries if s["section_id"] == sec_id), "")
            tags = ", ".join(spec_tags_map.get(sec_id, []))
            ws.append([sec_id, title, tags])
        if not extra_in_spec:
            ws.append(["None"])

        ws.append([])

        # Tags summary
        all_tags = {}
        for tags in spec_tags_map.values():
            for tag in tags:
                all_tags[tag] = all_tags.get(tag, 0) + 1

        ws.append(["Tag Summary"])
        ws.append(["Tag", "Count"])
        for cell in ws[ws.max_row]:
            cell.font = openpyxl.styles.Font(bold=True)
        if all_tags:
            for tag, count in sorted(all_tags.items()):
                ws.append([tag, count])
        else:
            ws.append(["No tags found"])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            adjusted_width = max_length + 2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        try:
            wb.save(self.output_file)
            logging.info(f"Validation report saved successfully.")
        except Exception as e:
            logging.error(f"Failed to save validation report: {e}")

    def run(self):
        if self.validate():
            self.write_report()


from pathlib import Path

if __name__ == "__main__":
    PROJECT_ROOT = Path(__file__).resolve().parents[2]
    toc_path = PROJECT_ROOT / "output" / "usb_pd_toc.jsonl"
    spec_path = PROJECT_ROOT / "output" / "usb_pd_spec.jsonl"
    output_path = PROJECT_ROOT / "output" / "validation_report.xlsx"

    validator = SectionValidator(toc_path, spec_path, output_path)
    validator.run()
