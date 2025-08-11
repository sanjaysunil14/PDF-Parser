import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class ValidationReportGenerator:
    def __init__(self):
        self.toc_data = []
        self.content_data = []
        self.metadata = {}
        
    def load_data(self):
        """Load data from JSONL files"""
        # Load TOC data
        try:
            with open("usb_pd_toc.jsonl", "r", encoding="utf-8") as f:
                for line in f:
                    if line.strip():
                        self.toc_data.append(json.loads(line))
        except FileNotFoundError:
            print("Warning: usb_pd_toc.jsonl not found")
        
        # Load content data
        try:
            with open("usb_pd_spec.jsonl", "r", encoding="utf-8") as f:
                for line in f:
                    if line.strip():
                        self.content_data.append(json.loads(line))
        except FileNotFoundError:
            print("Warning: usb_pd_spec.jsonl not found")
        
        # Load metadata
        try:
            with open("usb_pd_metadata.jsonl", "r", encoding="utf-8") as f:
                self.metadata = json.load(f)
        except FileNotFoundError:
            print("Warning: usb_pd_metadata.jsonl not found")
    
    def analyze_validation(self):
        """Analyze differences between TOC and parsed content"""
        validation_results = {
            "toc_sections": len(self.toc_data),
            "parsed_sections": len(self.content_data),
            "matches": 0,
            "mismatches": [],
            "missing_in_content": [],
            "extra_in_content": [],
            "order_errors": [],
            "gaps": [],
            "table_counts": self.count_tables(),
            "figure_counts": self.count_figures()
        }
        
        # Create lookup dictionaries
        toc_lookup = {entry["section_id"]: entry for entry in self.toc_data}
        content_lookup = {section["section_id"]: section for section in self.content_data}
        
        # Check matches and mismatches
        for toc_entry in self.toc_data:
            section_id = toc_entry["section_id"]
            if section_id in content_lookup:
                validation_results["matches"] += 1
                # Check for order errors or page mismatches
                content_section = content_lookup[section_id]
                if abs(toc_entry["page"] - content_section.get("page_start", toc_entry["page"])) > 2:
                    validation_results["order_errors"].append({
                        "section_id": section_id,
                        "toc_page": toc_entry["page"],
                        "content_page": content_section.get("page_start", "N/A"),
                        "difference": abs(toc_entry["page"] - content_section.get("page_start", toc_entry["page"]))
                    })
            else:
                validation_results["missing_in_content"].append({
                    "section_id": section_id,
                    "title": toc_entry["title"],
                    "page": toc_entry["page"]
                })
        
        # Check for extra sections in content
        for content_section in self.content_data:
            section_id = content_section["section_id"]
            if section_id not in toc_lookup:
                validation_results["extra_in_content"].append({
                    "section_id": section_id,
                    "title": content_section["title"],
                    "page": content_section.get("page_start", "N/A")
                })
        
        return validation_results
    
    def count_tables(self):
        """Count tables in content vs TOC references"""
        toc_table_count = 0
        content_table_count = 0
        
        # Count table references in TOC titles
        for entry in self.toc_data:
            if "table" in entry["title"].lower():
                toc_table_count += 1
        
        # Count actual tables found in content
        for section in self.content_data:
            content_table_count += len(section.get("tables", []))
        
        return {
            "toc_references": toc_table_count,
            "content_found": content_table_count,
            "metadata_count": self.metadata.get("total_tables", 0)
        }
    
    def count_figures(self):
        """Count figures in content vs TOC references"""
        toc_figure_count = 0
        content_figure_count = 0
        
        # Count figure references in TOC titles
        for entry in self.toc_data:
            if "figure" in entry["title"].lower():
                toc_figure_count += 1
        
        # Count actual figures found in content
        for section in self.content_data:
            content_figure_count += len(section.get("figures", []))
        
        return {
            "toc_references": toc_figure_count,
            "content_found": content_figure_count,
            "metadata_count": self.metadata.get("total_figures", 0)
        }
    
    def generate_excel_report(self, filename="usb_pd_validation_report.xlsx"):
        """Generate comprehensive validation report in Excel format"""
        validation_results = self.analyze_validation()
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # 1. Summary Sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet.append(["USB PD Specification Validation Report"])
        summary_sheet.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        summary_sheet.append([])
        
        # Overall statistics
        summary_sheet.append(["Metric", "Expected (TOC)", "Actual (Parsed)", "Status"])
        summary_sheet.append(["Total Sections", validation_results["toc_sections"], 
                             validation_results["parsed_sections"], 
                             "MATCH" if validation_results["toc_sections"] == validation_results["parsed_sections"] else "MISMATCH"])
        summary_sheet.append(["Matched Sections", "", validation_results["matches"], ""])
        summary_sheet.append(["Missing in Content", "", len(validation_results["missing_in_content"]), ""])
        summary_sheet.append(["Extra in Content", "", len(validation_results["extra_in_content"]), ""])
        summary_sheet.append(["Order Errors", "", len(validation_results["order_errors"]), ""])
        summary_sheet.append([])
        
        # Table and figure counts
        summary_sheet.append(["Tables (TOC References)", validation_results["table_counts"]["toc_references"], 
                             validation_results["table_counts"]["content_found"], ""])
        summary_sheet.append(["Figures (TOC References)", validation_results["figure_counts"]["toc_references"], 
                             validation_results["figure_counts"]["content_found"], ""])
        
        # Apply styles to summary sheet
        for row in range(1, summary_sheet.max_row + 1):
            for col in range(1, 5):
                cell = summary_sheet.cell(row=row, column=col)
                if row == 1:  # Title
                    cell.font = Font(bold=True, size=16)
                elif row == 4:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                elif col == 4 and cell.value == "MISMATCH":
                    cell.fill = error_fill
                elif col == 4 and cell.value == "MATCH":
                    cell.fill = success_fill
        
        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(filename)
        print(f"Validation report saved to {filename}")
        
        return validation_results

def main():
    generator = ValidationReportGenerator()
    generator.load_data()
    results = generator.generate_excel_report()
    
    print("\nValidation Summary:")
    print(f"TOC Sections: {results['toc_sections']}")
    print(f"Parsed Sections: {results['parsed_sections']}")
    print(f"Matches: {results['matches']}")
    print(f"Issues Found: {len(results['missing_in_content']) + len(results['extra_in_content']) + len(results['order_errors'])}")

if __name__ == "__main__":
    main()
